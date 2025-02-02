import datetime
import tkinter
from random import randint
from openpyxl.reader.excel import load_workbook
from tkinter import filedialog, messagebox as mbox

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from Utils import auxiliary_tools as tls
import settings
from app_windows.popup import Popup


class ExcelMixin:
    def __init__(self):
        self.wb = None

    def load_excel(self, path: str) -> None:
        """
        Загрузить рабочую книгу Excel по переданному пути.

        :param path: путь до выбранного файла Excel в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        """
        self.wb = load_workbook(path)

    def excel_open_file(self) -> str:
        """
        Открыть файл Excel через диалоговое окно и загрузить рабочую книгу wb

        :return: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        """
        file_path = filedialog.askopenfilename(title='Выберите файл Excel', filetypes=[('Все файлы', '*.xls*')])
        file_path = file_path.replace('/', '\\\\')              # Для корректной работы в Windows
        self.load_excel(file_path)                              # Загрузка базы из выбранного файла
        return file_path

    def excel_save_file_as(self) -> str:
        """
        Сохранить файл Excel как. Вызывает диалоговое окно с выбором папки сохранения и запросом имени файла.

        :return: str имя сохранённого файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\Saved_file.xlsx'
        """

        folder_selected = filedialog.askdirectory(title="Выберите папку для сохранения")
        if folder_selected:
            filename = filedialog.asksaveasfilename(            # Открываем диалог для ввода имени файла
                initialdir=folder_selected,
                title="Введите имя файла",
                defaultextension=".xlsx",
                filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
            )
            if filename:
                filename = filename.replace('/', '\\')          # Привет Винда
                self.wb.save(filename)
                return filename

    def excel_create_new_workbook_db(self) -> str | None:
        """
        Создание новой книги Excel со структурой базы - с системными листами + необходимым форматированием и формулами.

        Создаёт книгу с парой тестовых записей, запускает диалоговое окно с выбором папки сохранения и имени файла,
        загружает книгу.

        :return: str имя созданной книги (если успешно) или None - при возникновении ошибок.
        """
        try:
            # Создаём новую книгу и записываем её в self.wb для дальнейшей работы
            self.wb = Workbook()

            # Настройка первого листа
            ws = self.wb.active
            ws.title = 'Тема 1'
            ws.sheet_properties.tabColor = settings.EXCEL_SHEET_TAB_COLOR                   # Цвет ярлычка

            # Записываем и форматируем заголовки столбцов
            for col, value in settings.EXCEL_COLUMNS_WORD_SHEET.items():
                ws[f'{col}{1}'] = value['header']
                ws[f'{col}{1}'].font = Font(bold=True)
                ws.column_dimensions[col].width = value['width']                        # Устанавливаем ширину столбцов

            # Проставляем прогрессию ID
            index_first_row = settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA
            index_max_row = settings.EXCEL_AUTOINDEX_LIMIT
            for num in range(index_first_row, index_max_row + 1):
                ws[f'A{num}'] = num + 1 - index_first_row

            # Настроим перенос текста, выравнивание для столбцов, форматирование контекста.
            for col in settings.EXCEL_COLUMNS_WORD_SHEET.keys():
                for row in range(index_first_row, index_max_row):
                    cell = ws[f'{col}{row}']
                    if col != 'A':
                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    if col == 'E':
                        cell.font = Font(italic=True, color="064681")

            # Записываем тестовые строки с примером заполнения
            for num in range(1, 4):
                test_data = [f'word {num}', f'transcription {num}', f'Перевод - слово {num}',
                             f'Here is some example with word {num}.    And one more example with word 1 {num}.   '
                             f'{"test_wrapping " * 10}']
                for col, value in zip(['B', 'C', 'D', 'E'], test_data):
                    ws[f'{col}{index_first_row + num - 1}'] = value

            # Минимально простая самопроверка-тестирование внутри страницы
            if index_first_row > 5:
                cells = [settings.RAND_INDEX, settings.RAND_WORD, settings.RAND_WORD_TRANSCR, settings.RAND_WORD_TRANSL,
                         settings.RAND_WORD_CONTEXT]
                values = [settings.formula_rand_index, settings.formula_rand_word, settings.formula_rand_transcr,
                          settings.formula_rand_transl, settings.formula_rand_context]

                # Заполняем и настраиваем область самопроверки в Excel
                ws.freeze_panes = settings.EXCEL_WORD_SHEET_FROZEN_RANGE                        # Закрепляем диапазон
                ws[settings.RAND_TIP[0]] = settings.RAND_TIP[1]                         # Заполняем значения и формулы
                ws[settings.RAND_TIP[0]].font = Font(color=settings.EXCEL_TEXT_ADD_COLOR, bold=True)
                for ind, (cell, val) in enumerate(zip(cells, values)):
                    ws[cell] = val
                    if ind > 1:
                        ws[cell].font = Font(color=settings.EXCEL_TEXT_TIPS_COLOR)          # Форматируем подсказки

                for row in ws[settings.EXCEL_RANGE_BOTTOM_BORDER]:
                    for cell in row:
                        cell.border = Border(bottom=Side(border_style='thick', color='000000'))     # Чертим границу

            # Создаём лист со статистикой
            ws = self.wb.create_sheet(settings.EXCEL_STAT_LIST)
            self.wb._sheets.insert(0, self.wb._sheets.pop(self.wb._sheets.index(ws)))       # Перемещаем лист в начало

            # Создаём хедер таблицы, форматируем, выравниваем, задаём ширину столбцов
            for col, value in settings.EXCEL_COLUMNS_STAT_SHEET.items():
                ws[f'{col}1'] = value['header']
                ws[f'{col}1'].font = Font(bold=True)
                ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[col].width = value['width']

            # Форматирование отображения %
            for row in range(index_first_row + 1, index_max_row):   # Применяем формат для всех ячеек в столбце D
                ws[f'D{row}'].number_format = '0.0%'                # Форматируем как процент с 1 знаком после запятой

            # Заполняем тестовую строку
            test_data = [
                datetime.datetime(2025, 1, 23, 15, 40), 901, 927, 0.972, 'Тема 1', 2000, 1, settings.TEST_TYPES[0]]
            for col, value in zip(settings.EXCEL_COLUMNS_STAT_SHEET.keys(), test_data):
                ws[f'{col}2'] = value

            # Создаём лист с попытками
            ws = self.wb.create_sheet(settings.EXCEL_ATTEMPTS)
            self.wb._sheets.insert(1, self.wb._sheets.pop(self.wb._sheets.index(ws)))       # Перемещаем лист в начало

            # Создаём хедер таблицы, форматируем, выравниваем, задаём ширину столбцов
            for col, value in settings.EXCEL_COLUMNS_ATTEMPTS_SHEET.items():
                ws.column_dimensions[col].width = value['width']
                ws[f'{col}1'] = value['header']
                ws[f'{col}1'].font = Font(bold=True)
                ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

            # Устанавливаем третий лист как активный - чтобы не открывать книгу на системных
            self.wb.active = 2

            # Запускаем диалоговое окно пути сохранения и сохраняем новую книгу.
            return self.excel_save_file_as()

        except Exception as error:
            mbox.showerror('', f'Ошибка выполнения операции:\n{error}')
            return None

    def excel_create_sheet(self, sheet_name: str, excel_path: str) -> str:
        """
        Создание нового листа в книге Excel.
        Функция копирует существующий лист со всеми настройками форматирования и формулами, очищает его от значений
        и сохраняет под переданным именем.

        :param sheet_name: str название нового листа
        :param excel_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :return: str название рабочего файла 'This_file.xlsx'
        """

        # Создаём копию листа с готовой структурой и переименовываем её
        model_sheet = self.wb[self.excel_get_list_all_sheets()[0]]
        new_sheet = self.wb.copy_worksheet(model_sheet)
        new_sheet.title = sheet_name

        # Очищаем диапазон значений (оставляя форматирование и формулы)
        index_first_row = settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA
        index_max_row = settings.EXCEL_AUTOINDEX_LIMIT
        for row in new_sheet[f'B{index_first_row}:E{index_max_row}']:
            for cell in row:
                cell.value = None                                   # Очищаем только данные, форматирование сохраняется

        # Сохраняем книгу
        write_path = self.excel_render_filepath_to_writepath(excel_path)
        self.wb.save(write_path)
        wb_name = write_path.split('\\')[-1]
        return wb_name

    def excel_delete_sheet(self, sheet_name: str, excel_path: str) -> None:
        """
        Удаление листа Excel по названию

        :param sheet_name: str название листа
        :param excel_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :return: None
        """
        work_sheet = self.wb[sheet_name]                        # Определяем лист
        self.wb.remove(work_sheet)                              # Удаляем объект листа из книги
        write_path = self.excel_render_filepath_to_writepath(excel_path)
        self.wb.save(write_path)                                # Сохраняем изменения

    def excel_edit_sheet_name(self, old_sheet_name: str, new_sheet_name: str, excel_path: str) -> None:
        """
        Переименование листа Excel

        :param old_sheet_name: str текущее название листа
        :param new_sheet_name: str новое желаемое название листа
        :param excel_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :return:
        """
        work_sheet = self.wb[old_sheet_name]                    # Определяем лист
        work_sheet.title = new_sheet_name                       # Переименовываем лист
        write_path = self.excel_render_filepath_to_writepath(excel_path)
        self.wb.save(write_path)                                # Сохраняем изменения

    def excel_create_row(self, popup_object: Popup) -> None:
        """
        Создание новой записи перевода в книге Excel по данным из формы

        :param popup_object: объект всплывающего окна, в котором вызвана функция, с данными из форм
        :return: None
        """
        work_sheet = self.wb[popup_object.choosen_sheet_cmbbx]
        words_now = self._get_counter_rows_from_excel(excel_sheet=work_sheet)
        new_row_index = words_now + settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA    # ИНДЕКС строки под запись
        columns = ['A', 'B', 'C', 'D', 'E']                                     # Рабочие колонки
        cells = [words_now + 1, popup_object.word, popup_object.transcript, popup_object.translate,
                 popup_object.context]                                          # Все значения к записи
        for column, cell in zip(columns, cells):
            work_sheet[f'{column}{new_row_index}'] = cell

        write_path = self.excel_render_filepath_to_writepath(popup_object.master.excel_path)
        self.wb.save(write_path)                                                # Сохраняем запись

    def excel_delete_row(self, popup_object: Popup) -> bool:
        """
        Удаление строки из листа Excel из формы ввода данных + переиндексация столбца А после

        :param popup_object: объект всплывающего окна, в котором вызвана функция, с данными из форм
        :return: True при успехе, False при несовпадении данных и отмене действия
        """
        wb = self.wb
        work_sheet = wb[popup_object.given_choosen_sheet_cmbbx]
        index = self.excel_get_row_index_by_id(popup_object.given_row_id)
        if work_sheet[f'B{index}'].value != popup_object.given_word:
            return False                                    # Если ИНДЕКС не соответствует значению, не трогаем запись
        work_sheet.delete_rows(index)                       # удаление строки
        self.excel_get_new_ids_for_rows(work_sheet)         # Переиндексация строк в книге Excel после удаления
        write_path = self.excel_render_filepath_to_writepath(popup_object.master.excel_path)
        wb.save(write_path)
        return True

    def excel_delete_row_without_form(self, row: list, widget: tkinter.Frame) -> bool:
        """
        Удаление записи из листа без формы, по строке с данными из таблицы

        :param row: список с данными для отображения строки в таблице,  row:
                    [4, 'word', 'transcription', 'translate', 'context']
        :param widget: объект tkinter.Frame, из которого вызывается метод
        :return: bool: True при успехе, False при ошибке или отмене действия
        """
        wb = self.wb
        work_sheet = wb[widget.parent.choosen_excel_list]
        index = self.excel_get_row_index_by_id(row[0])
        if work_sheet[f'B{index}'].value != row[1]:
            mbox.showerror('Ошибка', 'Ошибка индексации!\nID строки не соответствует значению.\nПроверьте файл')
            return False                                    # Если ИНДЕКС не соответствует значению, не трогаем запись

        work_sheet.delete_rows(index)                       # удаление строки
        self.excel_get_new_ids_for_rows(work_sheet)         # Переиндексация строк в книге Excel после удаления
        write_path = self.excel_render_filepath_to_writepath(widget.parent.excel_path)
        wb.save(write_path)
        return True

    def excel_edit_row_data(self, popup_object: Popup) -> bool:
        """
        Редактирование строки из формы ввода данных

        :param popup_object: объект всплывающего окна, в котором вызвана функция, с данными из форм
        :return: True при успехе, False при ошибке и отмене действия
        """
        index = self.excel_get_row_index_by_id(popup_object.given_row_id)
        work_sheet = self.wb[popup_object.given_choosen_sheet_cmbbx]
        if work_sheet[f'B{index}'].value != popup_object.given_word:
            return False                                # Если ИНДЕКС не соответствует значению, не трогаем запись
        else:
            columns = ['B', 'C', 'D', 'E']
            new_data = [popup_object.word, popup_object.transcript, popup_object.translate, popup_object.context]
            for column, new_value in zip(columns, new_data):
                work_sheet[f'{column}{index}'] = new_value

            write_path = self.excel_render_filepath_to_writepath(popup_object.master.excel_path)
            self.wb.save(write_path)                    # Сохраняем изменения в книгу Excel
            return True

    def excel_get_row_index_by_id(self, row_id: int) -> int:
        """
        Расчёт ИНДЕКСа строки Excel по id записи

        :param row_id: int id записи на странице
        :return: int ИНДЕКС этой строки в файле Excel
        """
        return row_id + settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA - 1

    def excel_render_filepath_to_writepath(self, excel_file_path: str) -> str:
        """
        Правка системно определяемого пути к файлу Excel на PATH для записи в файл Excel (для Windows).
        Да, они разные. По одному типу записи он файл видит, по другому сохраняет.

        :param excel_file_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :return: str путь до того же файла, но в формате для записи C:\ Users\ I_AM_USER\ Desktop\ This_file.xlsx
        """
        write_path = excel_file_path.replace('\\\\', '\\')                  # Правим path файла для записи
        if 'xlsm' in write_path:                        # Не убиваем макросы! Если файл был .xlsm создастся копия .xlsx
            write_path = write_path.replace('xlsm', 'xlsx')
        return write_path

    def excel_get_list_all_sheets(self) -> list[str]:
        """
        Достаем из эксель список всех листов со словами (БЕЗ системных)

        :return: Список названий листов БЕЗ системных ['Лист1', 'Лист2', 'Лист3', ... 'Лист8', 'Лист9']
        """
        all_sheets = self.wb.sheetnames             # Все листы
        all_word_sheets = list(filter(lambda x: x not in settings.EXCEL_UTILS_SHEETS, all_sheets))
        return all_word_sheets                       # После фильтрации

    def excel_is_empty_sheet(self, sheet_name: str) -> bool:
        """
        Проверка листа на наличие записей

        :param sheet_name: str название листа
        :return: True если лист пуст (нет записей в ключевой ячейке), False если записи есть
        """
        """  """
        active_sheet = self.wb[sheet_name]
        if active_sheet[settings.EXCEL_KEY_CELL].value is None:
            return True
        return False

    def _get_sheet_from_excel(self, sheet: str = None):
        """
        Функция возвращает объект листа Excel по переданному str названию или генерирует случайный из всех листов книги,
        кроме системных и пустых.

        :param sheet: str название листа Excel
        :return: объект-лист Excel
        """
        all_word_sheets = self.excel_get_list_all_sheets()  # Все листы

        for sh in all_word_sheets.copy():                   # Исключаем листы без записей слов
            check = self.excel_is_empty_sheet(sh)
            if check:
                all_word_sheets.remove(sh)
                                            # Если передано название листа, то оно, если нет - рандомизатор названия
        list_name = sheet if sheet else all_word_sheets[randint(0, len(all_word_sheets) - 1)]
        target_list = self.wb[list_name]
        return target_list                                  # Объект листа Excel в рабочей книге по названию

    def _get_counter_rows_from_excel(self, excel_sheet, index_from: int = None) -> int:
        """
        Посчитать кол-во заполненных строк в выбранном листе Excel. Счёт по ключевому столбцу.

        :param excel_sheet: ОБЪЕКТ лист Excel
        :param index_from: c какой строки начинать счёт (например, для таблиц =1 - если ВСЕГО строк и =2 если записей)
        :return: int количество заполненных строк, начиная с указанного индекса строки
        """
        words_total = 0
        index_from = index_from if index_from else settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA

        for row in excel_sheet.iter_rows(
                min_row=index_from,
                min_col=settings.EXCEL_KEY_COLUMN[0],
                max_col=settings.EXCEL_KEY_COLUMN[0]):
            for cell in row:
                if cell.value is not None:
                    words_total += 1
        return words_total

    def excel_get_random_row(self, sheet: str = None) -> list[int | str]:
        """ Достать случайную строку с данными из Excel

        :sheet: str название листа в Excel для выбора именно с этого листа
        :return: список с данными значений строки в Excel + названием листа
                [ID_записи_int, 'word', 'transcript', 'translate', 'Here is context', 'sheet name']
        """
        work_list = self._get_sheet_from_excel(sheet)                           # Достаём лист Excel
        words_total = self._get_counter_rows_from_excel(work_list)
        rand_index = randint(settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA,  # Достаём рандомный индекс строки
                             words_total + settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA - 1)
                                                                                # Достаём рандомную строку
        random_row = next(work_list.iter_rows(min_row=rand_index, max_row=rand_index, min_col=settings.EXCEL_DATA_COLUMNS[0],
                                              max_col=settings.EXCEL_DATA_COLUMNS[1]))
        random_row = [cell.value for cell in random_row]                        # Делаем список со значениями
        tls.replace_non_breaking_space(random_row)                              # Форматируем строку для вывода
        tls.context_formatting(random_row)
        random_row.append(work_list.title)                                      # Добавляем название листа
        return random_row

    def excel_get_row_by_id_and_sheet(self, row: list) -> list[int | str]:
        """
        Достать из файла актуальные данные строки по id записи с заданного листа. Например, после редактирования.

        :param row: список с исходными данными строки:
                    [3, 'word', 'transcript', 'translate', 'Here is context', 'sheet name']
        :return: список с актуальными данными строки по данному индексу листа:
                [3, 'new_word', 'new_transcript', 'new_translate', 'Here is new_context', 'sheet name']
        """
        work_list = self.wb[row[5]]                                                 # Рабочий лист
        row_index_in_excel = row[0] + settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA - 1  # ИНДЕКС строки в Excel по id записи
        row_in_file = list(work_list.iter_rows(min_row=row_index_in_excel, max_row=row_index_in_excel,
                                               min_col=settings.EXCEL_DATA_COLUMNS[0],
                                               max_col=settings.EXCEL_DATA_COLUMNS[1]))
        result = list([cell.value for cell in row_in_file[0]])                      # Забираем новые данные
        result.append(row[5])                                                       # Добавляем лист в список
        return result

    def excel_get_all_data_from_sheet(self, sheet_name: str) -> list[list[int | str]]:
        """
        Получить данные всех строк листа Excel

        :param sheet_name: str название листа Excel
        :return: список со списками данных строк формата:
                [[5, 'sixfold (adv, adj)', 'fōld', 'шестикратный, в шесть раз.', ''], [...] ...]
        """
        wsh = self.wb[sheet_name]
        words_total = self._get_counter_rows_from_excel(wsh)
        data = []
        for row in wsh.iter_rows(
                min_row=settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA,
                max_row=words_total + settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA - 1,
                min_col=settings.EXCEL_DATA_COLUMNS[0],
                max_col=settings.EXCEL_DATA_COLUMNS[1]):
            row = [cell.value for cell in row]
            data.append(row)
        data.reverse()
        return data

    def excel_get_new_ids_for_rows(self, worksheet) -> None:
        """
        Переиндексация, присвоение новых id строкам в столбце А в Excel после удаления. БЕЗ сохранения файла

        :param worksheet: ОБЪЕКТ лист Excel
        :return: None
        """
        start_id = 1
        for cell in worksheet.iter_rows(min_row=settings.EXCEL_INDEX_FIRST_ROW_WITH_DATA, min_col=1, max_col=1):
            if cell[0].value is not None:
                cell[0].value = start_id
                start_id += 1
            else:
                break

    def excel_get_statistic_sheet_data(self) -> list[list[int | str]]:
        """
        Сформировать индексированный список строк статистики из Excel

        :return: список со списками данных отчётов статистики (reversed).
                [[9, datetime.datetime(2024, 12, 13, 0, 0), 483, 554, 87,1, ' - ', 'words_total',  'report_type'], ... ]
        """
        var_stat_list = self.wb[settings.EXCEL_STAT_LIST]                   # лист со статой
        data = []

        # Забираем данные из Excel
        for row in var_stat_list.iter_rows(min_row=2, max_col=8):
            row_values = [cell.value for cell in row if cell.value != None]
            if row_values:
                 data.append(row_values)
            else:
                break

        # Форматируем данные для вывода
        for index, row in enumerate(data):
            row.insert(0, row[6])                                       # report_id из column G встаёт первым в списке
            row[4] = round(row[4] * 100, 1)                             # Приводим дробь к % ответов
            row[1] = row[1].date()                                      # Приводим dt-объект к дате YYYY-MM-DD
            row[7] = row[8]                                             # Проставляем test_type

        return reversed(data)

    def excel_write_row_statistic(self, stat_data: tuple[int, int, int], topic: str, file_path: str,
                                  test_type: str) -> bool:
        """
        Запись строки-отчёта статистики в файл Excel на лист статистики

        :param stat_data: кортеж данных (right, wrong, attempts)
        :param topic: str выбранный лист или "-"
        :param file_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :param test_type: str название ТИПА тестирования в отчёте
        :return: True при успехе выполнения
        """
        work_sheet = self.wb[settings.EXCEL_STAT_LIST]

        # Собираем все данные для записи
        dt = datetime.datetime.now()
        right, wrong, attempts = stat_data
        total = 0
        all_sheets = self.excel_get_list_all_sheets()
        for sheet in all_sheets:
            res = self._get_counter_rows_from_excel(self.wb[sheet])
            total += res                                                        # Всего слов

        rows_total = self._get_counter_rows_from_excel(work_sheet, index_from=1)
        index_new_row = rows_total + 1                                          # Индекс строки для новой записи
        new_repo_id = work_sheet[f'G{rows_total}'].value + 1                    # ID последнего отчёта + 1

        # Записываем в цикле
        cells = [dt, right, attempts, right/attempts, topic, total, new_repo_id, test_type]
        for column, cell in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], cells):
            work_sheet[f'{column}{index_new_row}'] = cell

        self._set_report_id_to_attempts(new_repo_id, test_type)                 # Присвоить id отчёта попыткам

        write_path = self.excel_render_filepath_to_writepath(file_path)
        self.wb.save(write_path)                                                # Сохраняем запись
        return True

    def excel_delete_statistic_report_row(self, report_id: int, file_path) -> bool:
        """
        Удалить отчёт из листа статистики

        :param report_id: int ID отчёта статистики
        :param file_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :return: bool True при успехе, False при отмене операции
        """
        # Запрос подтверждения удаления
        answer = mbox.askyesno('Подтвердите удаление',
                               'Вы действительно хотите удалить запись?\nОтменить действие бедет невозможно')
        if answer == 'Нет' or answer is False:
            return False

        work_sheet = self.wb[settings.EXCEL_STAT_LIST]
        for index, row in enumerate(work_sheet.iter_rows()):
            if row[6].value == report_id:                           # Для строки с переданным id отчёта в column G
                work_sheet.delete_rows(index + 1)                   # Удаление строки из отчёта по ИНДЕКС

                attempt_sheet = self.wb[settings.EXCEL_ATTEMPTS]    # Deleted вместо id удалённого отчёта для attempts
                for attempt_row in attempt_sheet.iter_rows():
                    if attempt_row[4].value == report_id:
                        attempt_row[4].value = 'Deleted'

        write_path = self.excel_render_filepath_to_writepath(file_path)
        self.wb.save(write_path)                                    # Сохранить изменения
        return True

    def excel_write_attempt(self, excel_path: str, word: str, result: str, test_type: str) -> bool:
        """
        Запись попытки в лист Attempt

        :param excel_path: str путь до выбранного файла в формате 'C:\\Users\\I_AM_USER\\Desktop\\This_file.xlsx'
        :param word: слово/фраза
        :param result: "Верно"/"Неверно"
        :param test_type: str тип тестирования из атрибута окна
        :return: True при успехе
        """
        work_sheet = self.wb[settings.EXCEL_ATTEMPTS]                                       # Активируем лист попыток

        # Записываем данные
        rows_total = self._get_counter_rows_from_excel(work_sheet, index_from=1)
        index_new_row = rows_total + 1
        cells = [index_new_row - 1, datetime.datetime.now(), word, result, None, test_type]
        columns = [column for column in settings.EXCEL_COLUMNS_ATTEMPTS_SHEET.keys()]
        for column, cell in zip(columns, cells):
            work_sheet[f'{column}{index_new_row}'] = cell

        write_path = self.excel_render_filepath_to_writepath(excel_path)
        self.wb.save(write_path)
        return True

    def excel_get_current_attempts_info(self, test_type: str) -> tuple[int, int, int]:
        """
        Получить статистику верных, неверных ответов и всего попыток из листа Attempt.
        Возвращает все записи без присвоенного id отчёта

        :param test_type: str название ТИПА тестирования
        :return: кортеж (right_counter, wrong_counter, attempts)
        """
        work_sheet = self.wb[settings.EXCEL_ATTEMPTS]
        rows_total = self._get_counter_rows_from_excel(work_sheet, index_from=1)
        right_counter = 0
        wrong_counter = 0

        for row in work_sheet.iter_rows(min_row=2, max_row=rows_total, min_col=1, max_col=6):
            if row[4].value is None:
                if row[5].value == test_type:
                    if row[3].value == 'Верно':
                        right_counter += 1
                    elif row[3].value == 'Неверно':
                        wrong_counter += 1
        attempts = right_counter + wrong_counter

        return right_counter, wrong_counter, attempts

    def _set_report_id_to_attempts(self, report_id: int, test_type: str) -> None:
        """
        Проставить id отчёта всем попыткам из Attempts без id. Без сохранения файла.

        :param report_id: ID нового отчёта
        :param test_type: str тип тестирования
        :return: None
        """
        columns = ['A', 'B', 'C', 'D', 'E', 'F']
        work_sheet = self.wb[settings.EXCEL_ATTEMPTS]
        rows_total = self._get_counter_rows_from_excel(work_sheet, index_from=1)
        for row in work_sheet.iter_rows(min_row=2, max_row=rows_total, min_col=1, max_col=6):
            if row[4].value is None:
                if row[5].value == test_type:
                    row[4].value = report_id

    def excel_get_attempts_by_report_id(self, report_id: int) -> list[list[int | str]]:
        """
        Получить список с детализацией попыток по номеру отчёта

        :param report_id: id отчёта статистики
        :return: список списков, список всех попыток в отчёте с переданным id
                [[67, '2025-01-09 17:14:25', 'a supremacy', 'Верно', 39], [...], ... ],
        """
        work_sheet = self.wb[settings.EXCEL_ATTEMPTS]
        attempts_list = []
        for row in work_sheet.iter_rows(min_col=1, max_col=5):
            if row[4].value == report_id:
                row = [cell.value for cell in row]
                row[1] = row[1].strftime('%Y-%m-%d %H:%M:%S')
                attempts_list.append(row)
        return attempts_list
